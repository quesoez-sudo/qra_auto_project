"""
tests/diagnose_discrepancy.py
==============================
Evidence-gathering script for the ~20% discrepancy between Python and Excel.
Does NOT fix anything — only collects hard facts to identify root causes.

Questions to answer per event:
  ALL   : What labels are in the Z column?  Are we missing parameters?
  LPF   : Does xl_mat contain values > 1.0?  (kW/m² vs probability mismatch?)
  TOXIC : Are xl-nonzero cells at different spatial locations from py-nonzero?
           (directional plume vs circular Euclidean distance?)
  FF    : Same as TOXIC.
  CVE   : Are xl values always exactly {0,1} or are there intermediate values?
           Does the ignition point Python passes match what Excel uses?

Run:
    python tests/diagnose_discrepancy.py
"""
import sys, os, time
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import numpy as np
import openpyxl
import xlwings as xw
from dotenv import load_dotenv
load_dotenv()

from qra_engine import (
    QX, QY, SX, SY, XX, YY,
    read_thermal_scenarios, read_toxic_scenarios,
    read_explosion_scenarios, read_ff_scenarios,
    formula_thermal, formula_toxic, formula_explosion, formula_ff,
    dist_grid, _safe_float,
)
from qra_v6_engine import read_core_scenarios, IMPACT_CONFIG

_WORKSPACE  = os.getenv('WORKSPACE', os.getcwd())
_EXCEL_PATH = os.path.join(_WORKSPACE, 'KernelV0 (version 1) JetFireFormula.xlsx')

def _resolve_macro():
    for ext in ('.xlsm', '.xlsx'):
        p = os.path.join(_WORKSPACE, f'MacroQRAV6 (version 1){ext}')
        if os.path.exists(p):
            return p
    raise FileNotFoundError('MacroQRAV6 not found')

_MACRO_PATH = _resolve_macro()

# ── Column helpers ────────────────────────────────────────────────────────────

def _col_letter_to_num(s):
    n = 0
    for c in s.upper():
        n = n * 26 + (ord(c) - 64)
    return n

def _col_num_to_letter(n):
    s = ''
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def _make_range(col):
    end = _col_num_to_letter(_col_letter_to_num(col) + QX - 1)
    return f'{col}1:{end}{QY}'

# ── xlwings helpers ───────────────────────────────────────────────────────────

def _read_labels(ws, col):
    """Return {label: row_number} for all non-None cells in column."""
    vals = ws.range(f'{col}1:{col}50').value or []
    return {str(v).strip(): i + 1 for i, v in enumerate(vals) if v is not None}

def _set_params(ws, labels, val_col, **kwargs):
    for label, val in kwargs.items():
        r = labels.get(label)
        if r is not None:
            ws.range(f'{val_col}{r}').value = val

def _read_mat(ws, col):
    raw = ws.range(_make_range(col)).value
    if raw is None:
        return np.zeros((QY, QX))
    arr = np.zeros((QY, QX))
    for i, row in enumerate(raw):
        if row:
            for j, v in enumerate(row):
                arr[i, j] = float(v) if v is not None else 0.0
    return arr

def _weather(key):
    if key.endswith('Noche'): return 'Noche'
    if key.endswith('Dia'):   return 'Dia'
    return key.rsplit('/', 1)[-1]

# ── Core reader ───────────────────────────────────────────────────────────────

def _read_core(app, impact_id):
    wb = app.books.open(_MACRO_PATH)
    ws = wb.sheets['Core']
    ws.range('C2').value = impact_id
    time.sleep(1.0)
    app.api.Calculate()
    time.sleep(0.5)
    rows = read_core_scenarios(ws)
    wb.close()
    # Convert to qra_engine-style core dict
    core = {}
    for sc in rows:
        core[sc['key']] = {
            'x': sc['x'], 'y': sc['y'], 'size': sc['size'],
            'probs': {10: sc.get('prob', 0),  11: sc.get('p_jf', 0),
                      12: sc.get('p_lpf', 0), 13: sc.get('p_epf', 0),
                      14: sc.get('p_fb', 0),  15: sc.get('p_cve', 0),
                      16: sc.get('p_blv', 0), 17: sc.get('p_ff', 0)},
        }
    return core

# ── Analysis helpers ──────────────────────────────────────────────────────────

def show_labels(ws, label_col, val_col, ev_name):
    """Print every Z-column label and its current AA value."""
    labels = _read_labels(ws, label_col)
    print(f'\n  [{ev_name}] All labels in col {label_col}  (values in col {val_col}):')
    if not labels:
        print(f'    (no labels found in {label_col}1:{label_col}50)')
        return labels
    for lbl, r in sorted(labels.items(), key=lambda x: x[1]):
        cur = ws.range(f'{val_col}{r}').value
        print(f'    row {r:2d}  {lbl!r:<28} current={cur!r}')
    return labels


def xl_value_stats(xl_mat, ev_name, sc_key):
    """Report distribution of xl values — key for detecting kW vs probability."""
    flat = xl_mat.ravel()
    nz   = flat[flat != 0]
    print(f'\n  [{ev_name}] Excel matrix value distribution — {sc_key}')
    print(f'    nonzero cells : {len(nz):,} / {len(flat):,}')
    if len(nz) == 0:
        print('    (all zero — scenario probably not found in kernel)')
        return
    print(f'    min / max     : {nz.min():.6f}  /  {nz.max():.6f}')
    print(f'    mean / median : {nz.mean():.6f}  /  {np.median(nz):.6f}')
    pcts = np.percentile(nz, [10, 25, 50, 75, 90, 99])
    print(f'    p10…p99       : {" ".join(f"{p:.4f}" for p in pcts)}')
    uniq = np.unique(np.round(nz, 5))
    if len(uniq) <= 20:
        print(f'    unique values : {uniq.tolist()}')
    else:
        print(f'    unique count  : {len(uniq)}  (first 10: {uniq[:10].tolist()})')
    print(f'    cells > 1.0   : {(nz > 1.0).sum():,}   ← >0 means NOT probability')
    print(f'    cells = 1.0   : {(nz == 1.0).sum():,}')
    print(f'    cells 0<v<1   : {((nz > 0) & (nz < 1.0)).sum():,}')


def spatial_analysis(xl_mat, py_mat, sx, sy, ev_name, sc_key):
    """
    Decompose the nonzero regions into: overlap / xl_only / py_only.
    Checks if the mismatch is due to different LOCATIONS (shape) vs same locations
    but different values (formula).
    Also measures circularity vs elongation of xl nonzero region.
    """
    xl_nz = xl_mat != 0
    py_nz = py_mat != 0
    both  = xl_nz & py_nz
    xl_only = xl_nz & ~py_nz
    py_only = ~xl_nz & py_nz
    both_zero = ~xl_nz & ~py_nz

    print(f'\n  [{ev_name}] Spatial decomposition — {sc_key}')
    total = QX * QY
    print(f'    both zero     : {both_zero.sum():>8,}  ({100*both_zero.sum()/total:.1f}%)')
    print(f'    overlap (nz)  : {both.sum():>8,}  ({100*both.sum()/total:.1f}%)')
    print(f'    xl_only (nz)  : {xl_only.sum():>8,}  ({100*xl_only.sum()/total:.1f}%)')
    print(f'    py_only (nz)  : {py_only.sum():>8,}  ({100*py_only.sum()/total:.1f}%)')

    # Among overlap cells: how many values actually match?
    if both.sum() > 0:
        match_in_overlap = (np.abs(xl_mat[both] - py_mat[both]) < 1e-4).sum()
        print(f'    value match in overlap: {match_in_overlap:,} / {both.sum():,}  '
              f'({100*match_in_overlap/both.sum():.1f}%)')

    # Distance-from-source stats for xl_only cells
    d = dist_grid(sx, sy)
    if xl_only.sum() > 0:
        dists = d[xl_only]
        print(f'    xl_only dist from src: '
              f'min={dists.min():.1f}  max={dists.max():.1f}  mean={dists.mean():.1f} m')
    if py_only.sum() > 0:
        dists = d[py_only]
        print(f'    py_only dist from src: '
              f'min={dists.min():.1f}  max={dists.max():.1f}  mean={dists.mean():.1f} m')

    # Circularity check on xl nonzero region
    if xl_nz.sum() > 20:
        rows_idx, cols_idx = np.where(xl_nz)
        xc = XX[rows_idx, cols_idx]
        yc = YY[rows_idx, cols_idx]
        std_x, std_y = xc.std(), yc.std()
        ratio = max(std_x, std_y) / min(std_x, std_y) if min(std_x, std_y) > 0 else float('inf')
        cx, cy = xc.mean(), yc.mean()
        print(f'    xl footprint centroid : ({cx:.1f}, {cy:.1f}) m   source: ({sx:.1f}, {sy:.1f}) m')
        print(f'    xl footprint spread   : std_x={std_x:.1f}  std_y={std_y:.1f}  '
              f'axis-ratio={ratio:.3f}  (1.0 = perfect circle)')
    if py_nz.sum() > 20:
        rows_idx, cols_idx = np.where(py_nz)
        xc = XX[rows_idx, cols_idx]
        yc = YY[rows_idx, cols_idx]
        std_x, std_y = xc.std(), yc.std()
        ratio = max(std_x, std_y) / min(std_x, std_y) if min(std_x, std_y) > 0 else float('inf')
        cx, cy = xc.mean(), yc.mean()
        print(f'    py footprint centroid : ({cx:.1f}, {cy:.1f}) m')
        print(f'    py footprint spread   : std_x={std_x:.1f}  std_y={std_y:.1f}  '
              f'axis-ratio={ratio:.3f}')


def error_by_distance(xl_mat, py_mat, sx, sy, ev_name, sc_key, n_bins=12):
    """Profile the error as a function of distance from source."""
    d    = dist_grid(sx, sy)
    diff = np.abs(xl_mat - py_mat)
    xl_f = xl_mat.ravel()
    py_f = py_mat.ravel()
    d_f  = d.ravel()
    diff_f = diff.ravel()

    active = (xl_f != 0) | (py_f != 0)
    if not active.any():
        return
    d_max = np.percentile(d_f[active], 99)
    bins  = np.linspace(0, d_max, n_bins + 1)

    print(f'\n  [{ev_name}] Error by distance from source — {sc_key}')
    print(f'  {"dist range (m)":>20}  {"n":>7}  {"xl_mean":>10}  {"py_mean":>10}  '
          f'{"meanErr":>10}  {"match%":>8}')
    for i in range(n_bins):
        m = active & (d_f >= bins[i]) & (d_f < bins[i + 1])
        if not m.any():
            continue
        xl_m  = xl_f[m].mean()
        py_m  = py_f[m].mean()
        err_m = diff_f[m].mean()
        pct   = (diff_f[m] < 1e-4).mean() * 100
        print(f'  {bins[i]:8.1f}–{bins[i+1]:8.1f}  '
              f'{m.sum():>7d}  {xl_m:>10.4f}  {py_m:>10.4f}  {err_m:>10.4f}  {pct:>7.1f}%')


# ── Scenario lookup helpers ───────────────────────────────────────────────────

def _find_scenario_opx(ws_opx, key, reader_fn, core):
    """Return first scenario dict matching `key` from the openpyxl sheet."""
    scenarios = reader_fn(
        type('FakeWB', (), {'__getitem__': lambda self, k: ws_opx})(),
        core
    )
    return next((s for s in scenarios if s['key'] == key), None)


def _find_sc_raw(ws_opx, target_key, col_b=2):
    """Scan col B for target_key, return row number or None."""
    for r in range(2, ws_opx.max_row + 1):
        if ws_opx.cell(r, col_b).value == target_key:
            return r
    return None


# ── Event-specific scenario builders ─────────────────────────────────────────

def _build_thermal_sc(ws_opx, key, core):
    r = _find_sc_raw(ws_opx, key)
    if r is None: return None
    dists = [_safe_float(ws_opx.cell(r, c).value) for c in range(7, 17)]
    return dict(key=key, sx=core[key]['x'], sy=core[key]['y'], therm_dists=dists)

def _build_toxic_sc(ws_opx, key, core):
    r = _find_sc_raw(ws_opx, key)
    if r is None: return None
    blob = ws_opx.cell(r, 7).value
    if not blob: return None
    pairs = []
    for line in str(blob).strip().split('\n'):
        parts = line.strip().split(',')
        if len(parts) < 4: continue
        try:
            d, p = float(parts[0]), float(parts[3])
            if d >= 0 and p >= 0.01: pairs.append((d, p))
        except ValueError: continue
    if not pairs: return None
    pairs.sort(key=lambda t: t[0])
    return dict(key=key, sx=core[key]['x'], sy=core[key]['y'],
                tox_dists=np.array([t[0] for t in pairs]),
                tox_probs=np.array([t[1] for t in pairs]))

def _build_ff_sc(ws_opx, key, core):
    r = _find_sc_raw(ws_opx, key)
    if r is None: return None
    lfl  = _safe_float(ws_opx.cell(r, 5).value)
    lflf = _safe_float(ws_opx.cell(r, 6).value)
    if lfl is None or lflf is None: return None
    return dict(key=key, sx=core[key]['x'], sy=core[key]['y'],
                lfl_dist=lfl, lflf_dist=lflf)

def _build_cve_sc(ws_opx, key, core):
    r = _find_sc_raw(ws_opx, key)
    if r is None: return None
    dists = [_safe_float(ws_opx.cell(r, c).value) for c in range(10, 15)]
    ign_x = _safe_float(ws_opx.cell(r, 22).value)
    ign_y = _safe_float(ws_opx.cell(r, 23).value)
    sx = ign_x if ign_x is not None else core[key]['x']
    sy = ign_y if ign_y is not None else core[key]['y']
    return dict(key=key, sx=sx, sy=sy, ign_x=ign_x, ign_y=ign_y, exp_dists=dists)


# ── Main ─────────────────────────────────────────────────────────────────────

# Scenarios: (failing, passing) per event — from compare_excel results
TEST_CASES = {
    'TOXIC': {
        'impact_id': 16, 'sheet': 'ImpactToxMatrix',
        'label_col': 'Z', 'value_col': 'AA', 'mat_col': 'AB',
        'fail': 'D7204_gas/L1/ROG_FL30/H/5.5mDia',      # 5.7%  match
        'pass': 'D7301/ST1/REF_FL16/NA/1mNoche',          # 99.8% match
        'builder': _build_toxic_sc,
        'py_fn': lambda sc: formula_toxic(
            dist_grid(sc['sx'], sc['sy']), sc['tox_dists'], sc['tox_probs']),
    },
    'LPF': {
        'impact_id': 18, 'sheet': 'ImpactThermMatrix',
        'label_col': 'Z', 'value_col': 'AA', 'mat_col': 'AB',
        'fail': 'P7204AB(SUC)/L1/LHN_FL42/H/1mDia',      # 58.2% match
        'pass': 'D7301/L7/REF_FL16/H/1mDia',              # 98.5% match
        'builder': _build_thermal_sc,
        'py_fn': lambda sc: formula_thermal(
            dist_grid(sc['sx'], sc['sy']), sc['therm_dists']),
    },
    'FF': {
        'impact_id': 23, 'sheet': 'ImpactFFMatrix',
        'label_col': 'Z', 'value_col': 'AA', 'mat_col': 'AB',
        'fail': 'D7204_gas/L1/ROG_FL30/H/5.5mNoche',     # 45.3% match
        'pass': 'P7204AB/L7/LHN_FL42/H/1mNoche',          # 99.4% match
        'builder': _build_ff_sc,
        'py_fn': lambda sc: formula_ff(
            dist_grid(sc['sx'], sc['sy']), sc['lfl_dist'], sc['lflf_dist']),
    },
    'CVE': {
        'impact_id': 21, 'sheet': 'ImpactExpMatrix',
        'label_col': 'Z', 'value_col': 'AA', 'mat_col': 'AB',
        'fail': 'E7213_carcasa/150mm/HVN_FL50/H/1mNoche', # 15%   match
        'pass': 'D7308AB/150mm/REF_FL43/H/1mDia',         # 98.7% match
        'builder': _build_cve_sc,
        'py_fn': lambda sc: formula_explosion(
            dist_grid(sc['sx'], sc['sy']), sc['exp_dists']),
    },
}


def main():
    sep = '═' * 70

    # ── Phase 1: read Core for each event ────────────────────────────────────
    print('Phase 1 — reading Core from MacroQRAV6 per impact ID...')
    app_macro = xw.App(visible=False)
    cores = {}
    try:
        for ev, cfg in TEST_CASES.items():
            cores[ev] = _read_core(app_macro, cfg['impact_id'])
            print(f'  {ev}: {len(cores[ev])} Core keys')
    finally:
        app_macro.quit()

    # ── Phase 2: open KernelV0 openpyxl for data reading ─────────────────────
    print('\nPhase 2 — opening KernelV0 (openpyxl) for consequence data...')
    wb_opx = openpyxl.load_workbook(_EXCEL_PATH, data_only=True)

    # ── Phase 3: open KernelV0 xlwings, run scenarios ────────────────────────
    print('\nPhase 3 — running scenarios in KernelV0 (xlwings) & collecting evidence...')
    app_xl = xw.App(visible=False)
    try:
        wb_xl = app_xl.books.open(_EXCEL_PATH)

        for ev, cfg in TEST_CASES.items():
            print(f'\n{sep}')
            print(f'EVENT: {ev}   sheet={cfg["sheet"]}')
            print(sep)

            sheet     = cfg['sheet']
            lbl_col   = cfg['label_col']
            val_col   = cfg['value_col']
            mat_col   = cfg['mat_col']
            builder   = cfg['builder']
            py_fn     = cfg['py_fn']
            core      = cores[ev]
            ws_xl     = wb_xl.sheets[sheet]
            ws_opx    = wb_opx[sheet]

            # ── Step A: show ALL labels ───────────────────────────────────────
            labels = show_labels(ws_xl, lbl_col, val_col, ev)

            for label, sc_key in [('FAILING', cfg['fail']), ('PASSING', cfg['pass'])]:
                print(f'\n  ── {label}: {sc_key}')

                # Check scenario exists in Core
                if sc_key not in core:
                    print(f'  !! Key not found in Core for {ev} (impact_id={cfg["impact_id"]})')
                    continue

                sc = builder(ws_opx, sc_key, core)
                if sc is None:
                    print(f'  !! Could not build scenario from {sheet}')
                    continue

                sx, sy = sc['sx'], sc['sy']
                weather_str = _weather(sc_key)

                # ── Step B: set params in Excel and calculate ─────────────────
                _set_params(ws_xl, labels, val_col,
                            **{'Impact/Risk': 0, 'Scenario': sc_key,
                               'Weather': weather_str, 'X': sx, 'Y': sy,
                               'Probability': 0.0})
                app_xl.calculate()
                xl_mat = _read_mat(ws_xl, mat_col)
                py_mat = py_fn(sc)

                # ── Step C: value distribution (key for kW vs probability) ────
                xl_value_stats(xl_mat, ev, sc_key)

                # ── Step D: spatial decomposition (key for directional vs circular)
                spatial_analysis(xl_mat, py_mat, sx, sy, ev, sc_key)

                # ── Step E: error by distance ─────────────────────────────────
                error_by_distance(xl_mat, py_mat, sx, sy, ev, sc_key)

                # ── Step F: CVE-specific — check ignition coords ──────────────
                if ev == 'CVE':
                    print(f'\n  [{ev}] Ignition coords: ign_x={sc.get("ign_x")}, '
                          f'ign_y={sc.get("ign_y")}  (source: {core[sc_key]["x"]}, {core[sc_key]["y"]})')
                    # Read back what Excel has for X/Y after we set them
                    x_set = ws_xl.range(f'{val_col}{labels["X"]}').value if 'X' in labels else 'N/A'
                    y_set = ws_xl.range(f'{val_col}{labels["Y"]}').value if 'Y' in labels else 'N/A'
                    print(f'  [{ev}] X we wrote={sx:.2f}  Excel reads back X={x_set}')
                    print(f'  [{ev}] Y we wrote={sy:.2f}  Excel reads back Y={y_set}')

        wb_xl.close()
    finally:
        app_xl.quit()

    wb_opx.close()
    print(f'\n{sep}')
    print('Diagnostic complete.')


if __name__ == '__main__':
    main()
