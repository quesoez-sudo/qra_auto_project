import openpyxl

wb_f = openpyxl.load_workbook('KernelV0_v2_copy.xlsx', data_only=False)
wb_d = openpyxl.load_workbook('KernelV0_v2_copy.xlsx', data_only=True)

formula_col = {
    'ImpactThermMatrix': 'AB',
    'ImpactToxMatrix': 'AB',
    'ImpactFFMatrix': 'AB',
    'ImpactExpMatrix': 'AB',
    'ImpactJFMatrix': 'AL',
}

for sname, fcol in formula_col.items():
    ws_f = wb_f[sname]
    ws_d = wb_d[sname]
    print(f'\n{"="*60}')
    print(f'SHEET: {sname}')
    print(f'{"="*60}')

    # Max rows/cols used
    print(f'Dimensions: {ws_f.dimensions}')

    # Header row (row 1) - print all non-empty cells
    print('\n--- Row 1 headers ---')
    for cell in ws_f[1]:
        if cell.value is not None:
            print(f'  {cell.coordinate}: {repr(cell.value)[:120]}')

    # Formula cell
    fcell = fcol + '1'
    print(f'\n--- Formula at {fcell} ---')
    print(f'  FORMULA: {ws_f[fcell].value}')
    print(f'  VALUE:   {ws_d[fcell].value}')

    # Print first data row (row 2) to understand columns
    print('\n--- Row 2 (first data row) ---')
    for cell in ws_f[2]:
        if cell.value is not None:
            col_letter = cell.column_letter
            data_val = ws_d.cell(row=2, column=cell.column).value
            print(f'  {cell.coordinate}: formula={repr(str(cell.value))[:80]}  |  value={data_val}')

    # Print a few more rows to see data
    print('\n--- Rows 2-6 column A,B,C values ---')
    for r in range(2, 7):
        row_data = []
        for c in range(1, 12):
            v = ws_d.cell(row=r, column=c).value
            row_data.append(v)
        print(f'  Row {r}: {row_data}')
