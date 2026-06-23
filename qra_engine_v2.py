"""
QRA Engine v2
=============
Implements the new formulas from KernelV0 (version 2) New Formulas.xlsx.

Covered event types:
  Thermal  (ImpactThermMatrix AB1) — circular distance, thermal probit
  Flash Fire (ImpactFFMatrix AB1)  — step function on LFL / LFLF radius
  Toxic    (ImpactToxMatrix AB1)   — CSV-blob probability interpolation
  Jet Fire (ImpactJFMatrix AL1)    — ELLIPTICAL zones, 8 wind directions (NEW)

Explosion (ImpactExpMatrix) is deferred to a later iteration.

Source coordinates: grid centre (test mode).
Real source coords will come from the Core sheet of MacroQRAV6 later.

Output: output_v2/<event_type>/<scenario>_<weather>.csv
        output_v2/<event_type>_summary.csv   (key stats per scenario)
"""

import math
import csv
import numpy as np
from pathlib import Path
from scipy.special import erf
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
INPUT_FILE  = "KernelV0_v2_copy.xlsx"    # scenario + threshold data
CORE_FILE   = "KernelV0_v2_copy.xlsx"    # source coords + frequencies
CORE_SHEET  = "CoreExample"              # swap to "Core" + MacroQRAV6 for production
OUTPUT_DIR = Path("output_v2")

QX = 315
QY = 317
SX = 1.0698412698412698   # m / cell
SY = 1.069425             # m / cell

# kW/m² threshold labels (headers in ThermMatrix / JFMatrix rows 1)
KW_THRESHOLDS = np.array([1.6, 5.0, 7.3, 9.5, 12.5, 16.0, 20.9, 25.0, 30.0, 35.0])

# JF direction parameters (read from AK9/AK10/AK11 in the sheet)
JF_N_DIRS       = 8
JF_ANGLE_OFFSET = 0.0   # degrees
JF_ANGLE_STEP   = 0.0   # 0 -> uniform 360deg / N_dirs

# Toxic unit mode:
#   'meters' — blob col1 is in metres (confirmed: kernel distM is in m, blob col1 is same unit)
#   'km'     — divide distM by 1000 before matching (incorrect — kept for reference only)
TOXIC_UNIT_MODE = 'meters'

# ── Grid arrays ───────────────────────────────────────────────────────────────
# Pre-compute once; all formula functions take (dist_m, x_c, y_c) as arguments.

def _build_grids(src_x=None, src_y=None, qx=QX, qy=QY, sx=SX, sy=SY):
    """
    Returns:
        dist_m  (QY, QX)  Euclidean distance from source, metres
        x_c     (1,  QX)  x displacement in cell-width units from source  (for JF)
        y_c     (QY,  1)  y displacement in cell-height units from source (for JF)
    """
    cols = np.arange(1, qx + 1)
    rows = np.arange(1, qy + 1)

    # Cell-centre positions in metres
    x_m = sx * (cols - 0.5)                    # (QX,)
    y_m = sy * (qy - rows + 0.5)               # (QY,)  row 1 = northmost

    dist_m = np.sqrt(
        (x_m[None, :] - src_x) ** 2 +
        (y_m[:, None] - src_y) ** 2
    )                                           # (QY, QX)

    # Cell units from source — used by the JF ellipse formula
    src_col = src_x / sx
    src_row = src_y / sy
    x_c = ((cols - 0.5) - src_col)[None, :]   # (1, QX)
    y_c = ((qy - rows + 0.5) - src_row)[:, None]  # (QY, 1)

    return dist_m, x_c, y_c


# ── Thermal probit ────────────────────────────────────────────────────────────
def _thermal_prob(kw_mat, exp_time):
    """kW/m² array -> lethality probability array (same shape)."""
    p = np.zeros_like(kw_mat, dtype=float)
    m = (kw_mat > 0) & (kw_mat < 35.0)
    Y = -36.38 + 2.56 * np.log((1000.0 * kw_mat[m]) ** (4.0 / 3.0) * exp_time)
    p[m] = 0.5 * (1.0 + erf((Y - 5.0) / math.sqrt(2.0)))
    p[kw_mat >= 35.0] = 1.0
    return np.clip(p, 0.0, 1.0)


# ── THERMAL ───────────────────────────────────────────────────────────────────
def compute_thermal(dist_m, distances, exp_time=20.0, return_risk=False, frequency=1.0):
    """
    dist_m    : (QY, QX) distances in metres
    distances : array-like length 10 — each entry is the radial distance (m) at
                which kW/m² drops to KW_THRESHOLDS[i]. NaN / '' = threshold not reached.
    exp_time  : exposure seconds (from $AA$8)
    Returns   : kW/m² impact matrix (QY, QX), or risk matrix if return_risk=True
    """
    dv = np.array([float(x) if x not in ('', None) else np.nan for x in distances])
    valid = ~np.isnan(dv)
    dv_v = dv[valid]
    kw_v = KW_THRESHOLDS[valid]

    impact = np.zeros_like(dist_m, dtype=float)
    if len(dv_v) == 0:
        return impact

    # Inside every threshold -> max kW
    impact[dist_m <= dv_v[-1]] = kw_v[-1]

    # Linear interpolation between consecutive threshold distances
    for i in range(len(dv_v) - 1):
        mask = (dist_m > dv_v[i + 1]) & (dist_m <= dv_v[i])
        slope = (kw_v[i + 1] - kw_v[i]) / (dv_v[i + 1] - dv_v[i])
        impact[mask] = kw_v[i] + slope * (dist_m[mask] - dv_v[i])

    # Beyond outermost threshold -> 0 (already zero)

    if return_risk:
        return frequency * _thermal_prob(impact, exp_time)
    return impact


# ── FLASH FIRE ────────────────────────────────────────────────────────────────
def compute_ff(dist_m, lfl_r, lflf_r, return_risk=False, frequency=1.0):
    """
    lfl_r  : LFL radius (m)   — col E
    lflf_r : LFLF radius (m)  — col F
    Impact: 2 inside LFL, 1 between LFL-LFLF, 0 outside
    Risk  : frequency x (1 if inside LFL, else 0)
    """
    impact = np.zeros_like(dist_m, dtype=float)
    impact[dist_m < lflf_r] = 1.0
    impact[dist_m < lfl_r]  = 2.0
    if return_risk:
        return frequency * (impact == 2.0).astype(float)
    return impact


# ── TOXIC ─────────────────────────────────────────────────────────────────────
def _parse_tox_blob(blob_str, min_prob=0.01):
    """
    Parse the 5-column CSV blob from ImpactToxMatrix col G.

    Columns in the blob:
      col1: distance-like value (hypothesis: km) — descending, negative = upwind
      col2: concentration or dispersion parameter
      col3: log-like value
      col4: lethality probability (0-1)
      col5: secondary metric

    Filter: col1 >= 0  AND  col4 >= min_prob
    Returns (j_col, m_col) — sorted descending by col1 (distance).
    """
    rows = []
    for line in blob_str.strip().split('\n'):
        parts = line.split(',')
        if len(parts) < 5:
            continue
        try:
            vals = [float(p.strip()) for p in parts]
            rows.append(vals)
        except ValueError:
            continue

    if not rows:
        return np.array([]), np.array([])

    arr = np.array(rows)                         # (N, 5)
    keep = (arr[:, 0] >= 0) & (arr[:, 3] >= min_prob)
    filtered = arr[keep]

    if len(filtered) == 0:
        return np.array([]), np.array([])

    j_col = filtered[:, 0]   # col1 — distance-like, descending
    m_col = filtered[:, 3]   # col4 — lethality probability

    return j_col, m_col


def compute_toxic(dist_m, blob_str, return_risk=False, frequency=1.0,
                  unit_mode=TOXIC_UNIT_MODE):
    """
    dist_m   : (QY, QX) distances in metres
    blob_str : raw CSV blob from col G
    unit_mode: 'meters' -> compare distM directly to blob col1
               'km'     -> divide distM by 1000 before comparing (col1 treated as km)

    NOTE: The blob's col1 unit is ambiguous in this standalone kernel file.
    With unit_mode='km' the formula gives physically meaningful output when
    toxic clouds extend to hundreds of metres. Verify with real Core coords.
    """
    j_col, m_col = _parse_tox_blob(blob_str)
    if len(j_col) == 0:
        return np.zeros_like(dist_m, dtype=float)

    d = dist_m / 1000.0 if unit_mode == 'km' else dist_m

    j_max = j_col[0]    # largest (= farthest)
    j_min = j_col[-1]   # smallest positive

    # np.interp needs ascending x-coords; flip the descending j/m arrays
    j_asc = j_col[::-1]
    m_asc = m_col[::-1]

    out = np.interp(d, j_asc, m_asc, left=1.0, right=0.0)
    out = np.clip(out, 0.0, 1.0)

    if return_risk:
        return frequency * out
    return out


# ── JET FIRE ──────────────────────────────────────────────────────────────────
def compute_jf(x_c, y_c,
               dist_v, half_v, cent_v,
               n_dirs=JF_N_DIRS,
               angle_offset=JF_ANGLE_OFFSET,
               angle_step=JF_ANGLE_STEP,
               exp_time=20.0,
               return_risk=False,
               frequency=1.0):
    """
    Jet-fire elliptical impact formula.

    x_c, y_c : cell-unit displacement from source, shapes (1, QX) and (QY, 1)
    dist_v    : forward tip distances in cell units  (10 values, NaN if unused)
    half_v    : lateral half-width distances in cell units
    cent_v    : centre-from-source distances in cell units

    Algorithm (mirrors ImpactJFMatrix AL1 array formula):
      For each kW/m² threshold id (with valid data):
        ellipse centre  = c · (cos θ, sin θ)      [cell units]
        forward semi-ax = distV[id] - centerV[id]
        lateral semi-ax = halfWV[id]
        ellipse test    = (dx·cosθ + dy·sinθ)² / (d-c)² +
                          (dx·sinθ - dy·cosθ)² / halfW²  ≤ 1
        where dx = x - c·cosθ,  dy = y - c·sinθ
      impact[cell] = MAX(kW) over all (id, direction) where cell is inside ellipse
      Final: zero out cells beyond distV[0] (max forward reach) from source.

    Returns kW/m² impact matrix (QY, QX), or risk matrix if return_risk=True.
    """
    dv = np.array([float(x) if x not in ('', None) else np.nan for x in dist_v])
    hv = np.array([float(x) if x not in ('', None) else np.nan for x in half_v])
    cv = np.array([float(x) if x not in ('', None) else np.nan for x in cent_v])

    valid = ~(np.isnan(dv) | np.isnan(hv) | np.isnan(cv) |
              (dv == 0) | (hv == 0))
    dv_v = dv[valid]
    hv_v = hv[valid]
    cv_v = cv[valid]
    kw_v = KW_THRESHOLDS[valid]

    qy, qx = x_c.shape[1], y_c.shape[0]   # note: x_c (1,QX), y_c (QY,1)
    qy, qx = y_c.shape[0], x_c.shape[1]

    impact = np.zeros((qy, qx), dtype=float)

    if len(dv_v) == 0:
        return impact

    # Distance from source in cell units (for dmaxim outer mask)
    dist_cells = np.sqrt(x_c ** 2 + y_c ** 2)   # broadcast -> (QY, QX)
    d_maxim = dv_v[0]

    # Build angle list
    dirs = np.arange(n_dirs)
    if angle_step == 0:
        angles_deg = angle_offset + dirs / n_dirs * 360.0
    else:
        angles_deg = angle_offset + dirs * angle_step
    angles_rad = np.radians(angles_deg)

    for id_idx in range(len(kw_v)):
        d  = dv_v[id_idx]   # forward tip
        hw = hv_v[id_idx]   # lateral half-width
        c  = cv_v[id_idx]   # centre distance from source
        kw = kw_v[id_idx]

        arm = d - c          # forward semi-axis of ellipse
        if arm <= 0 or hw <= 0:
            continue

        a = 1.0 / arm ** 2   # aPower
        b = 1.0 / hw  ** 2   # bPower

        inside_any = np.zeros((qy, qx), dtype=bool)

        for theta in angles_rad:
            ct = math.cos(theta)
            st = math.sin(theta)

            # Vector from ellipse centre to each cell, in cell units
            dx = x_c - c * ct   # (1, QX) broadcasts with y_c (QY, 1) -> (QY, QX)
            dy = y_c - c * st

            # Ellipse quadratic form (= 1 on ellipse boundary)
            evals = (dx * ct + dy * st) ** 2 * a + \
                    (dx * st - dy * ct) ** 2 * b

            inside_any |= (evals <= 1.0)

        impact = np.maximum(impact, inside_any * kw)

    # Cells beyond max forward reach -> 0
    impact[dist_cells > d_maxim] = 0.0

    if return_risk:
        return frequency * _thermal_prob(impact, exp_time)
    return impact


# ── Sheet readers ─────────────────────────────────────────────────────────────
def _col(letter):
    """Column letter(s) -> 1-based column index."""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - 64)
    return n

def _v(ws, row, col_letter):
    return ws.cell(row=row, column=_col(col_letter)).value

def _flt(val):
    """Convert cell value to float or NaN."""
    if val is None or val == '':
        return np.nan
    try:
        return float(val)
    except (TypeError, ValueError):
        return np.nan


def read_therm(ws):
    scenarios = []
    for r in range(2, ws.max_row + 1):
        b = _v(ws, r, 'B')
        if not b:
            break
        distances = [_flt(_v(ws, r, c)) for c in ['F','G','H','I','J','K','L','M','N','O']]
        scenarios.append({
            'path_code': str(b),
            'scenario':  str(_v(ws, r, 'C')),
            'weather':   str(_v(ws, r, 'D')),
            'distances': distances,
        })
    return scenarios


def read_ff(ws):
    scenarios = []
    for r in range(2, ws.max_row + 1):
        b = _v(ws, r, 'B')
        if not b:
            break
        scenarios.append({
            'path_code': str(b),
            'scenario':  str(_v(ws, r, 'C')),
            'weather':   str(_v(ws, r, 'D')),
            'lfl_r':     _flt(_v(ws, r, 'E')),
            'lflf_r':    _flt(_v(ws, r, 'F')),
        })
    return scenarios


def read_tox(ws):
    scenarios = []
    for r in range(2, 5000):    # sheet has SEQUENCE(10000) filling H:H — stop at empty B
        b = _v(ws, r, 'B')
        if not b:
            break
        blob = _v(ws, r, 'G')
        scenarios.append({
            'path_code': str(b),
            'scenario':  str(_v(ws, r, 'C')),
            'weather':   str(_v(ws, r, 'D')),
            'blob':      str(blob) if blob else '',
        })
    return scenarios


def read_jf(ws):
    scenarios = []
    for r in range(2, ws.max_row + 1):
        b = _v(ws, r, 'B')
        if not b:
            break
        dist_v = [_flt(_v(ws, r, c)) for c in ['F','G','H','I','J','K','L','M','N','O']]
        half_v = [_flt(_v(ws, r, c)) for c in ['P','Q','R','S','T','U','V','W','X','Y']]
        cent_v = [_flt(_v(ws, r, c)) for c in ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']]
        scenarios.append({
            'path_code': str(b),
            'scenario':  str(_v(ws, r, 'C')),
            'weather':   str(_v(ws, r, 'D')),
            'flame_len': _flt(_v(ws, r, 'E')),
            'dist_v':    dist_v,
            'half_v':    half_v,
            'cent_v':    cent_v,
        })
    return scenarios


# ── Output helpers ────────────────────────────────────────────────────────────
def _safe_name(s):
    """Turn a scenario string into a safe filename stem."""
    return s.replace('/', '_').replace('\\', '_').replace(' ', '').replace('(', '').replace(')', '')


def _write_matrix_csv(path, matrix, header_lines, integer=False):
    """Write header comment lines + QYxQX numerical matrix to a CSV file."""
    with open(path, 'w', newline='') as f:
        w = csv.writer(f)
        for line in header_lines:
            w.writerow(['#' + line])
        if integer:
            for row in matrix:
                w.writerow([int(v) for v in row])
        else:
            for row in matrix:
                w.writerow([float(v) for v in row])


def _write_summary_csv(path, rows, fieldnames):
    with open(path, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def _radial_range(matrix, src_col, src_row, sx, sy, threshold):
    """Largest Euclidean distance (m) at which any cell has value >= threshold."""
    mask = matrix >= threshold
    if not mask.any():
        return 0.0
    row_idx, col_idx = np.where(mask)
    dx = sx * (col_idx + 1 - 0.5) - src_col
    dy = sy * (QY - row_idx - 0.5) - src_row
    return float(np.sqrt(dx**2 + dy**2).max())


def _to_risk(impact_mat, freq, event, exp_time=20.0):
    """Convert an impact matrix to a risk matrix (freq × lethality_probability)."""
    if event in ('thermal', 'jf'):
        return freq * _thermal_prob(impact_mat, exp_time)
    elif event == 'ff':
        # Only inside-LFL zone (value == 2) contributes to individual risk
        return freq * (impact_mat >= 2.0).astype(float)
    elif event == 'toxic':
        # Toxic impact is already a lethality probability
        return freq * impact_mat
    return np.zeros_like(impact_mat)


# ── Core loader ───────────────────────────────────────────────────────────────
# Frequency column indices (0-based) in MacroQRAV6 Core sheet
#   col A(0)=key, D(3)=X, E(4)=Y,
#   J(9)=P_TOXIC, K(10)=P_JF, L(11)=P_LPF, M(12)=P_EPF,
#   N(13)=P_FB,   O(14)=P_CVE, P(15)=P_BLV, Q(16)=P_FF, R(17)=P_LEXP
_CORE_FREQ_IDX = {
    'P_TOXIC': 9, 'P_JF': 10, 'P_LPF': 11, 'P_EPF': 12,
    'P_FB': 13, 'P_CVE': 14, 'P_BLV': 15, 'P_FF': 16, 'P_LEXP': 17,
}

def load_core(core_path=CORE_FILE, sheet_name=CORE_SHEET):
    """Return dict {key_scenario: {X, Y, P_TOXIC, P_JF, P_LPF, ...}} from Core sheet."""
    print(f"Loading Core from {core_path} / {sheet_name}...")
    wb = openpyxl.load_workbook(core_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    core = {}
    for row in ws.iter_rows(min_row=4, values_only=True):   # row 3 = header
        key = row[0]
        if not key:
            continue
        entry = {'X': row[3], 'Y': row[4]}
        for fname, idx in _CORE_FREQ_IDX.items():
            entry[fname] = row[idx] if idx < len(row) else None
        core[str(key)] = entry
    wb.close()
    print(f"  {len(core)} scenarios loaded from Core.")
    return core


def _lookup(core, key, freq_col):
    """Return (src_x, src_y, frequency) for a key, or None if missing."""
    entry = core.get(key)
    if entry is None:
        return None
    src_x = entry.get('X')
    src_y = entry.get('Y')
    freq  = entry.get(freq_col) or 0.0
    if src_x is None or src_y is None:
        return None
    return float(src_x), float(src_y), float(freq)


def _fmt_list(lst):
    """Format a list of floats/None for CSV header — compact, no huge decimals."""
    parts = []
    for v in lst:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            parts.append('None')
        else:
            parts.append(f'{float(v):.4g}')
    return '[' + ', '.join(parts) + ']'


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading kernel workbook...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

    core = load_core()   # reads CORE_FILE / CORE_SHEET constants at top

    OUTPUT_DIR.mkdir(exist_ok=True)
    for d in ('thermal', 'thermal_risk', 'ff', 'ff_risk', 'toxic', 'toxic_risk', 'jf', 'jf_risk'):
        (OUTPUT_DIR / d).mkdir(exist_ok=True)

    n_missing = {'thermal': 0, 'ff': 0, 'toxic': 0, 'jf': 0}

    # ── THERMAL ──────────────────────────────────────────────────────────────
    print("\n=== THERMAL (P_LPF from Core) ===")
    EXP_TIME = 20.0
    scenarios = read_therm(wb['ImpactThermMatrix'])
    summary = []

    for sc in scenarios:
        key = sc['path_code']
        coords = _lookup(core, key, 'P_LPF')
        if coords is None:
            print(f"  SKIP (no Core match): {key}")
            n_missing['thermal'] += 1
            continue
        src_x, src_y, freq = coords

        dist_m, _, _ = _build_grids(src_x, src_y)
        mat = compute_thermal(dist_m, sc['distances'], exp_time=EXP_TIME)

        name = _safe_name(key)
        _base_hdr = [
            f" KEY_SCENARIO  : {key}",
            f" EVENT         : thermal (ImpactThermMatrix / P_LPF)",
            f" QX            : {QX}",
            f" QY            : {QY}",
            f" SX_m_per_cell : {SX}",
            f" SY_m_per_cell : {SY}",
            f" SRC_X_m       : {src_x}",
            f" SRC_Y_m       : {src_y}",
            f" P_LPF         : {freq}",
            f" EXP_TIME_s    : {EXP_TIME}",
            f" KW_THRESHOLDS : {[float(x) for x in KW_THRESHOLDS]}",
            f" DISTANCES_m   : {_fmt_list(sc['distances'])}",
            f" GRID_NOTE     : row 1 = north edge (y = QY*SY - SY/2); col 1 = west edge",
        ]
        _write_matrix_csv(OUTPUT_DIR / 'thermal' / f"{name}.csv", mat,
                          _base_hdr + [" MATRIX_TYPE   : IMPACT  (kW/m2 interpolated)"])
        risk_mat = _to_risk(mat, freq, 'thermal', EXP_TIME)
        _write_matrix_csv(OUTPUT_DIR / 'thermal_risk' / f"{name}.csv", risk_mat,
                          _base_hdr + [" MATRIX_TYPE   : RISK  (P_LPF x lethality_prob)"])

        row_out = {'key_scenario': key, 'src_x_m': round(src_x, 4),
                   'src_y_m': round(src_y, 4), 'freq_P_LPF': freq}
        for kw in KW_THRESHOLDS:
            row_out[f'r_{kw}kW_m'] = round(_radial_range(mat, src_x, src_y, SX, SY, kw - 0.01), 1)
        summary.append(row_out)
        print(f"  {key}")

    _write_summary_csv(
        OUTPUT_DIR / 'thermal_summary.csv', summary,
        ['key_scenario', 'src_x_m', 'src_y_m', 'freq_P_LPF'] + [f'r_{kw}kW_m' for kw in KW_THRESHOLDS],
    )
    print(f"  -> output_v2/thermal  ({len(summary)} written, {n_missing['thermal']} skipped)")

    # ── FLASH FIRE ────────────────────────────────────────────────────────────
    print("\n=== FLASH FIRE (P_FF from Core) ===")
    scenarios = read_ff(wb['ImpactFFMatrix'])
    summary = []

    for sc in scenarios:
        key = sc['path_code']
        coords = _lookup(core, key, 'P_FF')
        if coords is None:
            print(f"  SKIP (no Core match): {key}")
            n_missing['ff'] += 1
            continue
        src_x, src_y, freq = coords

        dist_m, _, _ = _build_grids(src_x, src_y)
        mat = compute_ff(dist_m, sc['lfl_r'], sc['lflf_r'])

        name = _safe_name(key)
        _base_hdr = [
            f" KEY_SCENARIO  : {key}",
            f" EVENT         : flash fire (ImpactFFMatrix / P_FF)",
            f" QX            : {QX}",
            f" QY            : {QY}",
            f" SX_m_per_cell : {SX}",
            f" SY_m_per_cell : {SY}",
            f" SRC_X_m       : {src_x}",
            f" SRC_Y_m       : {src_y}",
            f" P_FF          : {freq}",
            f" LFL_R_m       : {sc['lfl_r']}",
            f" LFLF_R_m      : {sc['lflf_r']}",
            f" GRID_NOTE     : row 1 = north edge (y = QY*SY - SY/2); col 1 = west edge",
        ]
        _write_matrix_csv(OUTPUT_DIR / 'ff' / f"{name}.csv", mat,
                          _base_hdr + [" MATRIX_TYPE   : IMPACT  (0=safe 1=LFLF-zone 2=LFL-zone)"],
                          integer=True)
        risk_mat = _to_risk(mat, freq, 'ff')
        _write_matrix_csv(OUTPUT_DIR / 'ff_risk' / f"{name}.csv", risk_mat,
                          _base_hdr + [" MATRIX_TYPE   : RISK  (P_FF x 1_inside_LFL)"])

        lfl_comp  = round(_radial_range(mat, src_x, src_y, SX, SY, 1.9), 1)
        lflf_comp = round(_radial_range(mat, src_x, src_y, SX, SY, 0.9), 1)
        summary.append({
            'key_scenario': key, 'src_x_m': round(src_x, 4), 'src_y_m': round(src_y, 4),
            'freq_P_FF': freq,
            'lfl_r_input_m': sc['lfl_r'], 'lflf_r_input_m': sc['lflf_r'],
            'computed_lfl_r_m': lfl_comp, 'computed_lflf_r_m': lflf_comp,
        })
        print(f"  {key}")

    _write_summary_csv(
        OUTPUT_DIR / 'ff_summary.csv', summary,
        ['key_scenario','src_x_m','src_y_m','freq_P_FF',
         'lfl_r_input_m','lflf_r_input_m','computed_lfl_r_m','computed_lflf_r_m'],
    )
    print(f"  -> output_v2/ff  ({len(summary)} written, {n_missing['ff']} skipped)")

    # ── TOXIC ─────────────────────────────────────────────────────────────────
    print(f"\n=== TOXIC (P_TOXIC from Core, blob unit={TOXIC_UNIT_MODE!r}) ===")
    scenarios = read_tox(wb['ImpactToxMatrix'])
    summary = []

    for sc in scenarios:
        key = sc['path_code']
        coords = _lookup(core, key, 'P_TOXIC')
        if coords is None:
            print(f"  SKIP (no Core match): {key}")
            n_missing['toxic'] += 1
            continue
        src_x, src_y, freq = coords

        dist_m, _, _ = _build_grids(src_x, src_y)
        mat = compute_toxic(dist_m, sc['blob'])

        name = _safe_name(key)
        _base_hdr = [
            f" KEY_SCENARIO  : {key}",
            f" EVENT         : toxic (ImpactToxMatrix / P_TOXIC)",
            f" QX            : {QX}",
            f" QY            : {QY}",
            f" SX_m_per_cell : {SX}",
            f" SY_m_per_cell : {SY}",
            f" SRC_X_m       : {src_x}",
            f" SRC_Y_m       : {src_y}",
            f" P_TOXIC       : {freq}",
            f" BLOB_UNIT_MODE: {TOXIC_UNIT_MODE}",
            f" GRID_NOTE     : row 1 = north edge (y = QY*SY - SY/2); col 1 = west edge",
        ]
        _write_matrix_csv(OUTPUT_DIR / 'toxic' / f"{name}.csv", mat,
                          _base_hdr + [" MATRIX_TYPE   : IMPACT  (lethality probability 0-1)"])
        risk_mat = _to_risk(mat, freq, 'toxic')
        _write_matrix_csv(OUTPUT_DIR / 'toxic_risk' / f"{name}.csv", risk_mat,
                          _base_hdr + [" MATRIX_TYPE   : RISK  (P_TOXIC x lethality_prob)"])

        row_out = {'key_scenario': key, 'src_x_m': round(src_x, 4),
                   'src_y_m': round(src_y, 4), 'freq_P_TOXIC': freq}
        for p_thr in [0.01, 0.10, 0.50, 0.90]:
            row_out[f'r_{int(p_thr*100)}pct_m'] = round(
                _radial_range(mat, src_x, src_y, SX, SY, p_thr), 1)
        summary.append(row_out)
        print(f"  {key}")

    _write_summary_csv(
        OUTPUT_DIR / 'toxic_summary.csv', summary,
        ['key_scenario','src_x_m','src_y_m','freq_P_TOXIC',
         'r_1pct_m','r_10pct_m','r_50pct_m','r_90pct_m'],
    )
    print(f"  -> output_v2/toxic  ({len(summary)} written, {n_missing['toxic']} skipped)")

    # ── JET FIRE ──────────────────────────────────────────────────────────────
    print("\n=== JET FIRE (P_JF from Core) ===")
    scenarios = read_jf(wb['ImpactJFMatrix'])
    summary = []

    for sc in scenarios:
        key = sc['path_code']
        coords = _lookup(core, key, 'P_JF')
        if coords is None:
            print(f"  SKIP (no Core match): {key}")
            n_missing['jf'] += 1
            continue
        src_x, src_y, freq = coords

        _, x_c, y_c = _build_grids(src_x, src_y)
        mat = compute_jf(x_c, y_c, sc['dist_v'], sc['half_v'], sc['cent_v'],
                         exp_time=EXP_TIME)

        dv0 = sc['dist_v'][0]
        try:
            reach_m = float(dv0) * SX if dv0 not in ('', None) else 0.0
        except (TypeError, ValueError):
            reach_m = 0.0

        name = _safe_name(key)
        _base_hdr = [
            f" KEY_SCENARIO  : {key}",
            f" EVENT         : jet fire (ImpactJFMatrix / P_JF)",
            f" QX            : {QX}",
            f" QY            : {QY}",
            f" SX_m_per_cell : {SX}",
            f" SY_m_per_cell : {SY}",
            f" SRC_X_m       : {src_x}",
            f" SRC_Y_m       : {src_y}",
            f" P_JF          : {freq}",
            f" EXP_TIME_s    : {EXP_TIME}",
            f" FLAME_LEN_m   : {sc['flame_len']}",
            f" N_DIRS        : {JF_N_DIRS}  (uniform, every {360//JF_N_DIRS} deg)",
            f" KW_THRESHOLDS : {[float(x) for x in KW_THRESHOLDS]}",
            f" distV_cells   : {_fmt_list(sc['dist_v'])}",
            f" halfWV_cells  : {_fmt_list(sc['half_v'])}",
            f" centV_cells   : {_fmt_list(sc['cent_v'])}",
            f" distV_m       : {_fmt_list([v * SX if not (math.isnan(v) if isinstance(v, float) else False) else float('nan') for v in sc['dist_v']])}",
            f" GRID_NOTE     : row 1 = north edge (y = QY*SY - SY/2); col 1 = west edge",
        ]
        _write_matrix_csv(OUTPUT_DIR / 'jf' / f"{name}.csv", mat,
                          _base_hdr + [f" MATRIX_TYPE   : IMPACT  (kW/m2 MAX ellipse over {JF_N_DIRS} dirs x 10 thresholds)"])
        risk_mat = _to_risk(mat, freq, 'jf', EXP_TIME)
        _write_matrix_csv(OUTPUT_DIR / 'jf_risk' / f"{name}.csv", risk_mat,
                          _base_hdr + [" MATRIX_TYPE   : RISK  (P_JF x lethality_prob)"])

        row_out = {'key_scenario': key, 'src_x_m': round(src_x, 4),
                   'src_y_m': round(src_y, 4), 'freq_P_JF': freq,
                   'flame_len_m': sc['flame_len'],
                   'reach_1.6kW_m': round(reach_m, 1)}
        for kw in KW_THRESHOLDS:
            row_out[f'r_{kw}kW_m'] = round(
                _radial_range(mat, src_x, src_y, SX, SY, kw - 0.01), 1)
        summary.append(row_out)
        print(f"  {key}")

    _write_summary_csv(
        OUTPUT_DIR / 'jf_summary.csv', summary,
        ['key_scenario','src_x_m','src_y_m','freq_P_JF','flame_len_m','reach_1.6kW_m']
        + [f'r_{kw}kW_m' for kw in KW_THRESHOLDS],
    )
    print(f"  -> output_v2/jf  ({len(summary)} written, {n_missing['jf']} skipped)")

    print(f"\nAll done.  Results in: {OUTPUT_DIR.resolve()}")
    total_skip = sum(n_missing.values())
    if total_skip:
        print(f"  {total_skip} scenarios had no Core match (see SKIP lines above)")


if __name__ == '__main__':
    main()
