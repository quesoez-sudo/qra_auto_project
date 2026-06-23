import zipfile, xml.etree.ElementTree as ET

ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

with zipfile.ZipFile('KernelV0_v2_copy.xlsx') as z:
    xml_data = z.read('xl/worksheets/sheet7.xml')  # ImpactThermMatrix

root = ET.fromstring(xml_data)

print("=== ImpactThermMatrix row 1 formulas ===")
for row_el in root.findall('.//x:row', ns):
    if row_el.get('r') != '1':
        continue
    for c_el in row_el.findall('x:c', ns):
        ref = c_el.get('r')
        f_el = c_el.find('x:f', ns)
        if f_el is not None and f_el.text:
            ftype = f_el.get('t', 'normal')
            print(f"\nCell {ref} [type={ftype}]:")
            print(f_el.text)

# Also read params from AA column to see how srcX/srcY are referenced
print("\n\n=== AA column formulas (params) ===")
import openpyxl
wb = openpyxl.load_workbook('KernelV0_v2_copy.xlsx', data_only=False)
ws = wb['ImpactThermMatrix']
for r in [2,3,5,6,7,8,13,14,15,16,25]:
    cell = ws['AA' + str(r)]
    val_wb = openpyxl.load_workbook('KernelV0_v2_copy.xlsx', data_only=True)
    val = val_wb['ImpactThermMatrix']['AA' + str(r)].value
    print(f"  AA{r}: formula={repr(str(cell.value))[:80]}  value={repr(val)}")
    val_wb.close()
wb.close()
