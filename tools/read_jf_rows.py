"""Quick check of JF row structure - verify cols F-O / P-Y / Z-AI alignment."""
import openpyxl
wb = openpyxl.load_workbook('KernelV0_v2_copy.xlsx', data_only=True)
ws = wb['ImpactJFMatrix']

def col(letter):
    n = 0
    for c in letter:
        n = n * 26 + (ord(c) - ord('A') + 1)
    return n

# Print first 6 rows showing all 3 data bands
KW = [1.6, 5, 7.3, 9.5, 12.5, 16, 20.9, 25, 30, 35]
distV_cols  = ['F','G','H','I','J','K','L','M','N','O']
halfW_cols  = ['P','Q','R','S','T','U','V','W','X','Y']
centV_cols  = ['Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']

for r in range(2, 8):
    b = ws.cell(r, col('B')).value
    if not b:
        break
    sce = ws.cell(r, col('C')).value
    wth = ws.cell(r, col('D')).value
    fl  = ws.cell(r, col('E')).value
    dv = [ws.cell(r, col(c)).value for c in distV_cols]
    hv = [ws.cell(r, col(c)).value for c in halfW_cols]
    cv = [ws.cell(r, col(c)).value for c in centV_cols]
    print(f'\nRow {r}: {sce} | {wth} | flame={fl}')
    print(f'  distV  : {dv}')
    print(f'  halfWV : {hv}')
    print(f'  centV  : {cv}')

# Also check FF and Tox max data rows
for sname in ['ImpactFFMatrix', 'ImpactToxMatrix']:
    ws2 = wb[sname]
    count = 0
    for r in range(2, 10001):
        if ws2.cell(r, col('B')).value:
            count += 1
        else:
            break
    print(f'\n{sname}: {count} data rows')
