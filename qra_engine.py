"""
QRA Calculation Engine
======================
Computes 317×315 Impact and Risk matrices for 9 event types × 6 leak sizes.
Reads KernelV0 workbook (Core + ImpactXXMatrix sheets), outputs CSV files.
See engine_architecture.md for full design documentation.
"""
import numpy as np
import openpyxl
import os
import time
 
# ── Paths ────────────────────────────────────────────────────────────────────
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# TO RUN ON ANOTHER MACHINE: change ONLY _WORKSPACE to the folder that
# contains  "KernelV0 (version 1).xlsx"  on that machine.
# Everything else (EXCEL_PATH, OUTPUT_DIR, OUTPUT_EXCEL) is derived from it.
# Example:  _WORKSPACE = r'D:\Projects\QRA_tool'
# !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
_WORKSPACE   = r'C:\Users\herman.ramirez\OneDrive - Wood PLC\ODS\QRA_cod_project'
EXCEL_PATH   = _WORKSPACE + r'\KernelV0 (version 1).xlsx'
OUTPUT_DIR   = _WORKSPACE + r'\output'
OUTPUT_EXCEL = _WORKSPACE + r'\KernelV0 (version 1).xlsx'  # results written back in-place
 
# ── Grid (from Directions sheet rows 16-17) ──────────────────────────────────
QX = 315                        # columns (X axis)
QY = 317                        # rows (Y axis)
SX = 1.0698412698412698         # m per cell, X direction
SY = 1.069425                   # m per cell, Y direction
X_OFFSET = 0.0                  # X origin (m)
Y_OFFSET = 0.0                  # Y origin (m)
 
# Grid cell centers — computed once
_xc = X_OFFSET + SX * (np.arange(QX) + 0.5)   # (QX,)
_yc = Y_OFFSET + SY * (np.arange(QY) + 0.5)   # (QY,)
XX, YY = np.meshgrid(_xc, _yc)                 # (QY, QX)
 
SIZES = ['Total', 'S', 'M', 'L', 'XL', 'INST']
 
# ── Event configuration ──────────────────────────────────────────────────────
# prob_col: 1-indexed openpyxl column in Core sheet (J=10..Q=17)
EVENTS = [
    dict(name='TOXIC',    sheet='ImpactToxMatrix',   prob_col=10, formula='toxic'),
    dict(name='JF',       sheet='ImpactJFMatrix',    prob_col=11, formula='jf_ellipse'),
    dict(name='LPF',      sheet='ImpactThermMatrix', prob_col=12, formula='thermal'),
    dict(name='EPF',      sheet='ImpactThermMatrix', prob_col=13, formula='thermal'),
    dict(name='FB',       sheet='ImpactThermMatrix', prob_col=14, formula='thermal'),
    dict(name='CVE',      sheet='ImpactExpMatrix',   prob_col=15, formula='explosion', cve=True),
    dict(name='BLV',      sheet='ImpactExpMatrix',   prob_col=16, formula='explosion'),
    dict(name='FF',       sheet='ImpactFFMatrix',    prob_col=17, formula='ff'),
    dict(name='LATE_EXP', sheet='ImpactFFMatrix',    prob_col=None, formula='zero'),
]
 
# ── Thermal formula constants ─────────────────────────────────────────────────
# Radiation thresholds (kW/m²) read dynamically from ImpactThermMatrix row-1 headers
# (cols F-O = col indices 6-15).  The array below is a fallback used only when
# read_thermal_scenarios() cannot find header values.
THERM_THRESHOLDS = np.array([1.6, 5.0, 7.3, 9.5, 12.5, 16.0, 20.9, 25.0, 30.0, 35.0])
THERM_T_EXP = 20.0   # exposure time [s], from AA8
 
# ── Explosion formula constants ───────────────────────────────────────────────
# Overpressure thresholds (bar) → ImpactExpMatrix cols J-N (col indices 10-14)
EXP_THRESHOLDS = np.array([0.04, 0.1, 0.35, 0.5, 1.0])
EXP_LIM_OV1 = 0.1   # limitOV1 — AA15
EXP_LIM_OV2 = 0.3   # limitOV2 — AA16  (note: 0.3 not 0.35)
EXP_LIM_F1  = 0.0   # limitF1  — AA17
EXP_LIM_F2  = 1.0   # limitF2  — AA18
 
# ── Flash Fire formula constants ──────────────────────────────────────────────
FF_OUTSIDE    = 0.0   # limit1 (AA15)
FF_TRANSITION = 1.0   # limit2 (AA16)
FF_INSIDE_LFL = 2.0   # limit3 (AA17)

# ── Jet Fire formula constants ────────────────────────────────────────────────
# kW/m² thresholds → ImpactJFMatrix cols F-O (col indices 6-15)
JF_THRESHOLDS  = np.array([1.6, 5.0, 7.3, 9.5, 12.5, 16.0, 20.9, 25.0, 30.0, 35.0])
JF_DIRECTIONS  = 8      # number of equally-spaced jet-fire directions (AL25)
JF_ANGLE_OFFSET = 0.0   # starting angle in degrees (AL24)
JF_ANGLE_STEP  = 0      # 0 = equal spacing (360/n_dirs); >0 = fixed step [deg] (AL26)

 
# ── Toxic formula constants ───────────────────────────────────────────────────
TOX_MIN_PROB = 0.01   # AA15: minimum probability row to include from blob

# ── Wind direction (dominant, from WindMatrix sheet) ─────────────────────────
# Mathematical convention: 0° = East (+X), 90° = North (+Y).
# Overwritten at runtime by read_wind_direction(); default = North.
WIND_ANGLE_DEG = 90.0



# ── Helpers ───────────────────────────────────────────────────────────────────
def _safe_float(v):
    """Return float(v) or None for text / None values."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None
 
 
def _norm_cdf(x):
    """
    Standard normal CDF — vectorized, no scipy needed.
    Abramowitz & Stegun approximation, max error < 1.5e-7.
    """
    a1, a2, a3, a4, a5, p = (
        0.254829592, -0.284496736, 1.421413741,
        -1.453152027, 1.061405429, 0.3275911
    )
    x = np.asarray(x, dtype=float)
    sign = np.sign(x)
    xa = np.abs(x)
    t = 1.0 / (1.0 + p * xa)
    y = 1.0 - (((((a5 * t + a4) * t + a3) * t + a2) * t + a1) * t
                * np.exp(-xa * xa))
    return 0.5 * (1.0 + sign * y)
 
 
# ── Data readers ──────────────────────────────────────────────────────────────
 
def read_core(wb):
    """
    Read Core sheet into a dict keyed by ScenarioWeather string.
    Returns: { key: {x, y, size, probs:{col_idx: float}} }
    Core data starts at row 2 (with header at row 2?).  Skip rows where key=None.
    """
    ws = wb['Core']
    core = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value       # col A: ScenarioWeather
        if not key:
            continue
        x    = _safe_float(ws.cell(r, 4).value)   # col D: X [m]
        y    = _safe_float(ws.cell(r, 5).value)   # col E: Y [m]
        size = ws.cell(r, 7).value or ''           # col G: S/M/L/XL/INST
        if x is None or y is None:
            continue
        probs = {}
        for col in range(10, 18):   # J=10 (P_TOXIC) … Q=17 (P_FF)
            v = ws.cell(r, col).value
            probs[col] = float(v) if v else 0.0
        core[key] = dict(x=x, y=y, size=str(size).strip(), probs=probs)
    return core
 
 
def read_ff_scenarios(wb, core):
    """
    ImpactFFMatrix columns: B=path key, C=scenario name, E=LFL_dist[m], F=LFL_frac_dist[m].
    X/Y fall back to Core coordinates.
    Skip rows not in Core.
    """
    ws = wb['ImpactFFMatrix']
    scenarios = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 2).value
        if not key or key not in core:
            continue
        lfl_d  = _safe_float(ws.cell(r, 5).value)   # col E
        lflf_d = _safe_float(ws.cell(r, 6).value)   # col F
        if lfl_d is None or lflf_d is None:
            continue
        sx = core[key]['x']
        sy = core[key]['y']
        scenarios.append(dict(
            key=key, sx=sx, sy=sy,
            lfl_dist=lfl_d, lflf_dist=lflf_d,
            size=core[key]['size'], probs=core[key]['probs'],
        ))
    return scenarios
 
 
def read_thermal_scenarios(wb, core):
    """
    ImpactThermMatrix columns: B=key, F-O=radiation distances[m] for 10 thresholds.
    Row-1 headers in cols F-O hold the kW/m² values (1.6, 5.0, … 35.0).
    Some distance cells may contain 'Not reached…' text → stored as None.
    """
    ws = wb['ImpactThermMatrix']
    # Read kW/m² threshold values dynamically from row-1 headers (cols F-O = 6-15)
    thresholds = []
    for c in range(6, 16):
        v = _safe_float(ws.cell(1, c).value)
        if v is not None:
            thresholds.append(v)
    if not thresholds:
        thresholds = THERM_THRESHOLDS.tolist()   # fallback to module constant
    thresholds = np.array(thresholds)

    scenarios = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 2).value
        if not key or key not in core:
            continue
        dists = [_safe_float(ws.cell(r, c).value) for c in range(6, 16)]   # cols F-O
        scenarios.append(dict(
            key=key,
            sx=core[key]['x'], sy=core[key]['y'],
            therm_dists=dists,
            therm_thresholds=thresholds,
            size=core[key]['size'], probs=core[key]['probs'],
        ))
    return scenarios
 
 
def read_jf_scenarios(wb, core):
    """
    ImpactJFMatrix columns:
      B=key, F:O=far-tip distances (10 thresholds, cell units),
      P:Y=half-widths (semi-minor axis, cell units),
      Z:AI=center distances (cell units).
    """
    ws = wb['ImpactJFMatrix']
    scenarios = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 2).value
        if not key or key not in core:
            continue
        dist_vals   = [_safe_float(ws.cell(r, c).value) for c in range(6,  16)]
        halfW_vals  = [_safe_float(ws.cell(r, c).value) for c in range(16, 26)]
        center_vals = [_safe_float(ws.cell(r, c).value) for c in range(26, 36)]
        scenarios.append(dict(
            key=key,
            sx=core[key]['x'], sy=core[key]['y'],
            dist_vals=dist_vals,
            halfW_vals=halfW_vals,
            center_vals=center_vals,
            size=core[key]['size'], probs=core[key]['probs'],
        ))
    return scenarios
 
 
def read_explosion_scenarios(wb, core):
    """
    ImpactExpMatrix columns: B=key, J-N=overpressure distances[m], V=ign_X, W=ign_Y.
    """
    ws = wb['ImpactExpMatrix']
    scenarios = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 2).value
        if not key or key not in core:
            continue
        dists  = [_safe_float(ws.cell(r, c).value) for c in range(10, 15)]  # cols J-N
        ign_x  = _safe_float(ws.cell(r, 22).value)   # col V
        ign_y  = _safe_float(ws.cell(r, 23).value)   # col W
        scenarios.append(dict(
            key=key,
            sx=core[key]['x'], sy=core[key]['y'],
            ign_x=ign_x, ign_y=ign_y,
            exp_dists=dists,
            size=core[key]['size'], probs=core[key]['probs'],
        ))
    return scenarios
 
 
def read_toxic_scenarios(wb, core):
    """
    ImpactToxMatrix columns: B=key, G=CSV blob with dispersion profile.
    Blob format per line: distance[m], dose, log_val, probability, other
    Filters: distance >= 0 AND probability >= TOX_MIN_PROB.
    Sorted ascending by distance for np.interp.
    """
    ws = wb['ImpactToxMatrix']
    scenarios = []
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 2).value
        if not key or key not in core:
            continue
        blob = ws.cell(r, 7).value   # col G: CSV blob
        if not blob:
            continue
        pairs = []
        for line in str(blob).strip().split('\n'):
            parts = line.strip().split(',')
            if len(parts) < 4:
                continue
            try:
                d = float(parts[0])
                p = float(parts[3])
                if d >= 0 and p >= TOX_MIN_PROB:
                    pairs.append((d, p))
            except ValueError:
                continue
        if not pairs:
            continue
        pairs.sort(key=lambda t: t[0])   # ascending distance
        tox_dists = np.array([t[0] for t in pairs])
        tox_probs = np.array([t[1] for t in pairs])
        scenarios.append(dict(
            key=key,
            sx=core[key]['x'], sy=core[key]['y'],
            tox_dists=tox_dists, tox_probs=tox_probs,
            size=core[key]['size'], probs=core[key]['probs'],
        ))
    return scenarios
 
 
# ── Formula implementations ───────────────────────────────────────────────────
 
def dist_grid(sx, sy):
    """Euclidean distance (m) from point (sx, sy) to every grid cell → (QY, QX)."""
    return np.sqrt((XX - sx) ** 2 + (YY - sy) ** 2)


def read_wind_direction(wb):
    """
    Read dominant wind direction from WindMatrix sheet.

    WindMatrix layout:
      Row 1 – direction names (W, WNW, NW, …)
      Row 2 – angles in degrees (math convention: 0=East, 90=North)
      Row 3 – Dia   (daytime) frequencies
      Row 4 – Noche (nighttime) frequencies

    Returns the angle (float, degrees) of the direction with the highest
    combined Dia+Noche frequency.  Falls back to 90.0 (North) if the
    sheet is absent or unreadable.
    """
    if 'WindMatrix' not in wb.sheetnames:
        return 90.0
    ws = wb['WindMatrix']
    angles, totals = [], []
    for c in range(3, ws.max_column + 1):
        angle = _safe_float(ws.cell(2, c).value)
        dia   = _safe_float(ws.cell(3, c).value) or 0.0
        noche = _safe_float(ws.cell(4, c).value) or 0.0
        if angle is not None:
            angles.append(angle)
            totals.append(dia + noche)
    if not angles:
        return 90.0
    # WindMatrix angles are the direction FROM WHICH wind blows (meteorological FROM
    # convention).  Add 180° to get the downwind direction (where the wind goes).
    from_angle = float(angles[totals.index(max(totals))])
    return (from_angle + 180.0) % 360.0


def downwind_dist(sx, sy):
    """
    Signed downwind distance (m) from source (sx, sy) to every grid cell.
    Positive = downwind of source, negative = upwind.
    Uses WIND_ANGLE_DEG set by read_wind_direction().
    """
    angle_rad = np.radians(WIND_ANGLE_DEG)
    return (XX - sx) * np.cos(angle_rad) + (YY - sy) * np.sin(angle_rad)


def formula_ff(dist, lfl_dist, lflf_dist):
    """
    Flash Fire impact: step function on distance.
    Returns (QY, QX) array with values 0 / FF_TRANSITION / FF_INSIDE_LFL.
    """
    result = np.full(dist.shape, FF_OUTSIDE)
    result[dist <= lflf_dist] = FF_TRANSITION
    result[dist <= lfl_dist]  = FF_INSIDE_LFL
    return result
 
 
def formula_jf(sx, sy, dist_vals, halfW_vals, center_vals):
    """
    Jet Fire directional ellipse impact formula.

    For each of JF_DIRECTIONS equally-spaced jet angles, builds an ellipse for
    each kW/m² threshold (cols F:O, P:Y, Z:AI of ImpactJFMatrix).  A grid cell
    is assigned the highest threshold kW/m² value for which it falls inside at
    least one directional ellipse.

    All spatial quantities are in cell units (distances stored in ImpactJFMatrix
    are in cell units, consistent with the Excel formula dividing by SX/SY).

    Returns (QY, QX) array of kW/m² values (0.0 or one of JF_THRESHOLDS).
    """
    # Excel rounds cell-unit offsets to nearest integer (ROUND(...,0)).
    # Match that behaviour so boundary cells agree with the kernel output.
    x_rel = np.round((XX - sx) / SX)   # (QY, QX) integer cell-unit X offset
    y_rel = np.round((YY - sy) / SY)   # (QY, QX) integer cell-unit Y offset

    if JF_ANGLE_STEP == 0:
        angles_deg = JF_ANGLE_OFFSET + np.arange(JF_DIRECTIONS) * 360.0 / JF_DIRECTIONS
    else:
        angles_deg = JF_ANGLE_OFFSET + np.arange(JF_DIRECTIONS) * JF_ANGLE_STEP

    ct = np.cos(np.radians(angles_deg))   # (n_dirs,)
    st = np.sin(np.radians(angles_deg))   # (n_dirs,)

    x_exp = x_rel[..., np.newaxis]   # (QY, QX, 1) for broadcasting with n_dirs
    y_exp = y_rel[..., np.newaxis]

    result = np.zeros((QY, QX))

    for i in range(len(JF_THRESHOLDS)):
        dist   = dist_vals[i]
        halfW  = halfW_vals[i]
        center = center_vals[i]

        if dist is None or halfW is None or center is None:
            continue
        if dist <= 0 or halfW <= 0:
            continue

        a = dist - center   # semi-major axis
        b = halfW           # semi-minor axis
        c = center          # center distance from source

        if a <= 0:
            continue

        imp = JF_THRESHOLDS[i]

        # Projection onto jet axis and perpendicular for all directions at once
        proj_along = x_exp * ct + y_exp * st - c   # (QY, QX, n_dirs)
        proj_perp  = x_exp * st - y_exp * ct       # (QY, QX, n_dirs)

        eVals = (proj_along / a) ** 2 + (proj_perp / b) ** 2   # (QY, QX, n_dirs)

        inside_any = np.any(eVals <= 1.0, axis=2)   # (QY, QX) bool
        result = np.where(inside_any, np.maximum(result, imp), result)

    return result


def formula_thermal(dist, therm_dists, therm_thresholds=None):
    """
    Thermal radiation impact: return kW/m² at each grid cell.

    Pool fires (LPF/EPF/FB) radiate omnidirectionally — dist must be the
    Euclidean distance from the source (dist_grid), NOT a downwind projection.

    dist             – radial distance grid (m) from source; always ≥ 0.
    therm_dists      – list of radii (m) for each kW/m² threshold (may be None).
    therm_thresholds – 1-D array of kW/m² values matching therm_dists order.
                       Defaults to module-level THERM_THRESHOLDS if not supplied.

    Returns (QY, QX) array of kW/m² (0.0 outside all thresholds).
    """
    if therm_thresholds is None:
        therm_thresholds = THERM_THRESHOLDS

    valid_idx = [i for i, d in enumerate(therm_dists) if d is not None and d > 0]
    if not valid_idx:
        return np.zeros(dist.shape)

    kw_vals   = np.asarray(therm_thresholds)[valid_idx]
    dist_vals = np.array([therm_dists[i] for i in valid_idx])

    xp = dist_vals[::-1]   # ascending distances (small dist = high kW first)
    fp = kw_vals[::-1]     # descending kW

    return np.interp(
        dist, xp, fp,
        left=fp[0],    # inside innermost threshold → cap at max kW/m²
        right=0.0      # beyond outermost threshold → 0
    )
 
 
def formula_explosion(dist, exp_dists):
    """
    Explosion overpressure impact: interpolate bar pressure from threshold distances.

    Matches Excel's ImpactExpMatrix AB kernel (impact mode):
      impactM = linear interp(dist -> bar) using distV=cols J-N, impactV=[0.04..1.0 bar]
      - dist > MAX(distV) -> 0 bar
      - dist < MIN(distV) -> highest bar threshold (~1.0 bar)
      - otherwise         -> linearly interpolated bar pressure

    exp_dists: list of 5 values [dist_0.04bar, dist_0.1bar, dist_0.35bar, dist_0.5bar, dist_1.0bar].
    Distances are descending (J=farthest/lowest bar, N=closest/highest bar).

    Returns (QY, QX) array of bar pressure (0.0 to ~1.0).
    """
    valid_idx = [i for i, d in enumerate(exp_dists) if d is not None and d > 0]
    if not valid_idx:
        return np.zeros(dist.shape)

    bar_vals  = EXP_THRESHOLDS[valid_idx]
    dist_vals = np.array([exp_dists[i] for i in valid_idx])

    xp = dist_vals[::-1]   # ascending distances (small dist = high pressure)
    fp = bar_vals[::-1]    # descending bar

    return np.interp(dist, xp, fp, left=fp[0], right=0.0)
 
 
def formula_toxic(dist, tox_dists, tox_probs):
    """
    Toxic impact: interpolate fatality probability from the dispersion profile.

    The toxic cloud disperses radially (Excel uses a circle centred at the
    source), so dist must be the Euclidean distance from the source
    (dist_grid), NOT a signed downwind projection.

    dist      – radial distance grid (m); always ≥ 0 when using dist_grid.
    tox_dists – 1D ascending distances [m] where prob ≥ TOX_MIN_PROB.
    tox_probs – matching probability values.

    left = tox_probs[0]: cells closer than the first blob entry inherit the
    threshold probability (fills the centre of the circle instead of leaving
    a zero-probability hole caused by the blob's near-source dilution zone).

    Returns (QY, QX) probability array in [0, 1].
    """
    if len(tox_dists) == 0:
        return np.zeros(dist.shape)
    result = np.interp(
        dist, tox_dists, tox_probs,
        left=tox_probs[0],   # inside innermost entry → extend threshold prob inward
        right=0.0             # beyond max radius → 0
    )
    return np.clip(result, 0.0, 1.0)
 
 
# ── Main computation ──────────────────────────────────────────────────────────
 
def run_event(event, scenarios):
    """
    Accumulate Impact and Risk matrices for one event across all 6 size filters.
 
    Args:
        event: dict from EVENTS list
        scenarios: list of scenario dicts for this event's formula type
 
    Returns:
        dict: { size_label → (impact_matrix (QY,QX), risk_matrix (QY,QX)) }
    """
    formula  = event['formula']
    prob_col = event.get('prob_col')
    is_cve   = event.get('cve', False)
 
    # Pre-allocate output matrices for all 6 sizes
    impact_mats = {sz: np.zeros((QY, QX)) for sz in SIZES}
    risk_mats   = {sz: np.zeros((QY, QX)) for sz in SIZES}
 
    for sc in scenarios:
        sc_size = sc['size']
        prob = sc['probs'].get(prob_col, 0.0) if prob_col else 0.0
 
        # Source / ignition coordinates
        if is_cve:
            if sc['ign_x'] is None or sc['ign_y'] is None:
                continue   # CVE requires ignition point; skip if missing
            sx, sy = sc['ign_x'], sc['ign_y']
        else:
            sx, sy = sc['sx'], sc['sy']

        # ── Distance grid ──────────────────────────────────────────────────────
        # All consequence radii (toxic, thermal, ff, explosion) are measured
        # from the source/ignition point and applied as RADIAL distances:
        #
        #   toxic   – circle centred at source; formula_toxic uses dist_grid.
        #   thermal – omnidirectional pool/fireball radiation; dist_grid.
        #   ff      – LFL circle centred at source; dist_grid.
        #   explosion – radial overpressure from ignition point; dist_grid.
        #   jf_ellipse – handled inside formula_jf (directional ellipses).
        #
        # downwind_dist() is retained for any future directional model but is
        # not used in the standard event dispatch below.
        d = dist_grid(sx, sy)

        # Compute impact for this scenario
        if formula == 'thermal':
            cell_imp = formula_thermal(d, sc['therm_dists'], sc['therm_thresholds'])
        elif formula == 'explosion':
            cell_imp = formula_explosion(d, sc['exp_dists'])
        elif formula == 'toxic':
            cell_imp = formula_toxic(d, sc['tox_dists'], sc['tox_probs'])
        elif formula == 'ff':
            cell_imp = formula_ff(d, sc['lfl_dist'], sc['lflf_dist'])
        elif formula == 'jf_ellipse':
            cell_imp = formula_jf(sx, sy, sc['dist_vals'], sc['halfW_vals'], sc['center_vals'])
        else:
            continue   # 'zero' formula → skip
 
        risk_contrib = cell_imp * prob   # 0 if prob == 0
 
        # Accumulate into Total and into the matching size bucket
        impact_mats['Total'] += cell_imp
        risk_mats['Total']   += risk_contrib
 
        if sc_size in SIZES:
            impact_mats[sc_size] += cell_imp
            risk_mats[sc_size]   += risk_contrib
 
    return {sz: (impact_mats[sz], risk_mats[sz]) for sz in SIZES}
 
 
def save_csv(matrix, path):
    """Save 2D numpy array as CSV (scientific notation, 6 sig figs)."""
    np.savetxt(path, matrix, delimiter=',', fmt='%.6e')
 
 
# ── Excel output helpers ──────────────────────────────────────────────────────
 
def _col_letter_to_num(letters):
    """Convert Excel column letters to 1-indexed integer. 'A'→1, 'LC'→315."""
    num = 0
    for c in letters.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num
 
 
def _parse_range_start(range_str):
    """
    Parse a range string such as '$LD$318:$XF$634' and return the
    1-indexed (start_row, start_col) of the top-left cell.
    """
    clean = range_str.replace('$', '')
    start_cell = clean.split(':')[0]          # e.g. 'LD318'
    col_str = ''.join(c for c in start_cell if c.isalpha())
    row_str = ''.join(c for c in start_cell if c.isdigit())
    return int(row_str), _col_letter_to_num(col_str)
 
 
def read_directions_ranges(wb):
    """
    Read result-placement ranges from Directions sheet (columns E-J, rows 2-10).
    Returns: { event_name → { size_label → range_str } }
 
    Column mapping in Directions:
        E (col 5) = Total,  F (col 6) = S,  G (col 7) = M,
        H (col 8) = L,      I (col 9) = XL, J (col 10) = INST
    Row mapping follows the same order as the EVENTS list (rows 2-10).
    """
    ws = wb['Directions']
    size_col = {'Total': 5, 'S': 6, 'M': 7, 'L': 8, 'XL': 9, 'INST': 10}
    result = {}
    for i, ev in enumerate(EVENTS):
        row = 2 + i    # Directions data starts at row 2
        result[ev['name']] = {
            sz: str(ws.cell(row, col).value)
            for sz, col in size_col.items()
            if ws.cell(row, col).value
        }
    return result
 
 
def write_to_excel(output_path, all_results, directions_ranges):
    """
    Write all computed matrices to ImpactMatrix0 and RiskMatrix0 sheets
    inside the kernel workbook at output_path.
 
    - Loads the workbook WITHOUT data_only so all other sheet formulas are preserved.
    - Deletes and recreates ImpactMatrix0 and RiskMatrix0 (clean slate).
    - Writes EVERY cell including 0.0 so no cell is blank.
    """
    print('\nOpening kernel workbook for writing: %s' % output_path)
    t0 = time.time()
 
    # Load without data_only → formulas in other sheets are preserved on save
    wb = openpyxl.load_workbook(output_path)
    print('  Loaded in %.1fs. Sheets: %s' % (time.time() - t0, wb.sheetnames))
 
    # Clean slate: delete then recreate result sheets
    for sheet_name in ('ImpactMatrix0', 'RiskMatrix0'):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        wb.create_sheet(sheet_name)
 
    ws_imp  = wb['ImpactMatrix0']
    ws_risk = wb['RiskMatrix0']
 
    n_combos = len(EVENTS) * len(SIZES)
    combo    = 0
    total_blocks = 0
 
    for event in EVENTS:
        name         = event['name']
        ev_ranges    = directions_ranges.get(name, {})
        size_results = all_results.get(name, {})
 
        for sz in SIZES:
            combo += 1
            range_str = ev_ranges.get(sz)
            if not range_str:
                print('  [%d/%d] %-10s %-6s — no Directions range, skipping.'
                      % (combo, n_combos, name, sz))
                continue
 
            imp_mat, risk_mat = size_results.get(
                sz, (np.zeros((QY, QX)), np.zeros((QY, QX))))
            start_row, start_col = _parse_range_start(range_str)
 
            # Convert to Python lists once — faster element access than numpy
            imp_list  = imp_mat.tolist()
            risk_list = risk_mat.tolist()
 
            # Write every cell (including 0.0) — no blanks in result sheets
            for ri in range(QY):
                r        = start_row + ri
                imp_row  = imp_list[ri]
                risk_row = risk_list[ri]
                for ci in range(QX):
                    c = start_col + ci
                    ws_imp.cell(row=r,  column=c, value=imp_row[ci])
                    ws_risk.cell(row=r, column=c, value=risk_row[ci])
 
            total_blocks += 1
            print('  [%d/%d] %-10s %-6s → row %-5d col %-5d  (%d×%d cells)'
                  % (combo, n_combos, name, sz, start_row, start_col, QY, QX))
 
    total_cells = total_blocks * QY * QX
    print('Writing %d cells per sheet. Saving workbook (this will take a few minutes)...'
          % total_cells)
    wb.save(output_path)
    elapsed = time.time() - t0
    print('Done. %.1f min  |  %d cells written to each result sheet.'
          % (elapsed / 60, total_cells))
 
 
# ── Entry point ───────────────────────────────────────────────────────────────
 
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    t0 = time.time()
 
    print('Loading workbook: %s' % EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    global WIND_ANGLE_DEG
    WIND_ANGLE_DEG = read_wind_direction(wb)
    print('Wind direction  : %.1f° (math convention: 0=E, 90=N) from WindMatrix'
          % WIND_ANGLE_DEG)

    print('Reading Core sheet...')
    core = read_core(wb)
    print('  %d scenarios with coordinates loaded from Core.' % len(core))
 
    print('Reading impact sheets...')
    ff_scen    = read_ff_scenarios(wb, core)
    therm_scen = read_thermal_scenarios(wb, core)
    exp_scen   = read_explosion_scenarios(wb, core)
    tox_scen   = read_toxic_scenarios(wb, core)
    jf_scen    = read_jf_scenarios(wb, core) if 'ImpactJFMatrix' in wb.sheetnames else []
    print('  FF=%d  Thermal=%d  Explosion=%d  Toxic=%d  JF=%d  scenarios matched to Core'
          % (len(ff_scen), len(therm_scen), len(exp_scen), len(tox_scen), len(jf_scen)))
 
    scenario_map = {
        'ff':         ff_scen,
        'thermal':    therm_scen,
        'explosion':  exp_scen,
        'toxic':      tox_scen,
        'jf_ellipse': jf_scen,
        'zero':       [],
    }
 
    print('\nGrid: %d×%d cells, SX=%.4f m/cell, SY=%.4f m/cell' % (QX, QY, SX, SY))
    print('JF formula      : directions=%d  angle_offset=%.1f deg  angle_step=%s  thresholds=%s'
          % (JF_DIRECTIONS, JF_ANGLE_OFFSET,
             ('equal-spacing (360/n)' if JF_ANGLE_STEP == 0 else '%.1f deg' % JF_ANGLE_STEP),
             JF_THRESHOLDS.tolist()))
    print('Output: %s\n' % OUTPUT_DIR)
 
    n_events  = len(EVENTS)
    all_results = {}   # { event_name → { size → (impact_mat, risk_mat) } }
 
    for ei, event in enumerate(EVENTS, 1):
        name      = event['name']
        formula   = event['formula']
        scenarios = scenario_map[formula]
        t_ev      = time.time()
 
        print('[%d/%d] %s  (formula=%s, scenarios=%d)'
              % (ei, n_events, name, formula, len(scenarios)))
 
        results = run_event(event, scenarios)
        all_results[name] = results
 
        for sz in SIZES:
            imp, risk = results[sz]
            save_csv(imp,  os.path.join(OUTPUT_DIR, 'impact_%s_%s.csv' % (name, sz)))
            save_csv(risk, os.path.join(OUTPUT_DIR, 'risk_%s_%s.csv'   % (name, sz)))
 
        imp_max  = results['Total'][0].max()
        risk_max = results['Total'][1].max()
        print('  Done (%.1fs) | impact_Total max=%.4f | risk_Total max=%.4e'
              % (time.time() - t_ev, imp_max, risk_max))
 
    csv_elapsed = time.time() - t0
    print('\n%d CSV files written in %.1fs  →  %s'
          % (n_events * len(SIZES) * 2, csv_elapsed, OUTPUT_DIR))
 
    # ── Write results to Excel ─────────────────────────────────────────────
    directions_ranges = read_directions_ranges(wb)
    write_to_excel(OUTPUT_EXCEL, all_results, directions_ranges)
 
    print('\nTotal elapsed: %.1fs' % (time.time() - t0))
 
 
if __name__ == '__main__':
    main()
 