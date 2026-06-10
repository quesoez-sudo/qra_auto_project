"""
Scenario Audit — QRA Engine Diagnostic
=======================================

PURPOSE
-------
This script diagnoses why the Python engine produces lower impact values than
Excel's native computation.  The most common cause is missing scenario
contributions: Python silently skips scenarios whose key does not appear in Core.

SECTION A — Inventory Audit (always runs)
  For each event, reads every row from its impact sheet and reports:
    - Total rows in the sheet
    - Matched to Core (will be processed by the engine)
    - Skipped (key not in Core) + full list of skipped keys
    - Size distribution of matched scenarios
    - Probability column stats (min, max, non-zero count, zero count)

  USE THIS FIRST.  If skipped counts are high, that is why values are low.

SECTION B — Per-Scenario Contribution Log (controlled by RUN_CONTRIBUTION_LOG)
  For each matched scenario per event, runs the impact formula and records:
    Key, Size, Prob, MaxImpact, SumImpact, NonZeroCells
  Output is printed as a table and saved to output/scenario_audit_B.csv.

  USE THIS IF Section A is clean but the mismatch persists.
  Controls:
    RUN_CONTRIBUTION_LOG = True   — enable Section B (slower, runs all formulas)
    EVENTS_FILTER = ['TOXIC']     — limit to specific events; [] = all events

USAGE
-----
  1. Set _WORKSPACE to the folder containing "KernelV0 (version 1).xlsx"
  2. python tests/scenario_audit.py
  3. Check Section A for high skipped counts
  4. Set RUN_CONTRIBUTION_LOG = True and re-run if Section A is clean
"""

import os
import sys
import csv
import numpy as np
import openpyxl

# ── path so we can import from the project root ───────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from qra_engine import (
    QX, QY, SX, SY, XX, YY, EVENTS, SIZES,
    JF_THRESHOLDS, JF_DIRECTIONS, JF_ANGLE_OFFSET, JF_ANGLE_STEP,
    THERM_THRESHOLDS, THERM_T_EXP,
    EXP_THRESHOLDS, EXP_LIM_OV2, EXP_LIM_OV1,
    FF_OUTSIDE, FF_TRANSITION, FF_INSIDE_LFL, TOX_MIN_PROB,
    _safe_float,
    read_core, read_ff_scenarios, read_thermal_scenarios,
    read_explosion_scenarios, read_toxic_scenarios, read_jf_scenarios,
    formula_jf, formula_thermal, formula_explosion, formula_ff, formula_toxic,
    dist_grid,
)

# ── Configuration — edit these before running ─────────────────────────────────
_WORKSPACE           = r'C:\Users\herman.ramirez\OneDrive - Wood PLC\ODS\QRA_cod_project'
EXCEL_PATH           = os.path.join(_WORKSPACE, 'KernelV0 (version 1).xlsx')
OUTPUT_DIR           = os.path.join(_WORKSPACE, 'output')
RUN_CONTRIBUTION_LOG = False   # set True to run Section B (runs all formulas)
EVENTS_FILTER        = []      # e.g. ['TOXIC', 'LPF']; empty = all events

# ── Reader map: formula type → reader function ────────────────────────────────
_READER = {
    'toxic':      read_toxic_scenarios,
    'jf_ellipse': read_jf_scenarios,
    'thermal':    read_thermal_scenarios,
    'explosion':  read_explosion_scenarios,
    'ff':         read_ff_scenarios,
    'zero':       None,
}


# ── Section A ─────────────────────────────────────────────────────────────────

def run_section_a(wb, core, events):
    sep  = '─' * 72
    wide = '═' * 72
    print(f'\n{wide}')
    print('SECTION A — SCENARIO INVENTORY AUDIT')
    print(wide)

    for ev in events:
        name       = ev['name']
        sheet_name = ev['sheet']
        prob_col   = ev.get('prob_col')

        if sheet_name not in wb.sheetnames:
            print(f'\n[{name}]  sheet "{sheet_name}" NOT FOUND in workbook — skipping')
            continue

        ws         = wb[sheet_name]
        total_rows = 0
        matched    = []
        skipped    = []

        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 2).value   # col B = scenario key
            if not key:
                continue
            total_rows += 1
            if str(key) in core:
                matched.append(str(key))
            else:
                skipped.append(str(key))

        # Size distribution for matched scenarios
        size_dist = {}
        for key in matched:
            sz = core[key]['size'] or 'unknown'
            size_dist[sz] = size_dist.get(sz, 0) + 1

        # Probability stats for matched scenarios
        prob_stats = None
        if prob_col:
            probs     = [core[key]['probs'].get(prob_col, 0.0) for key in matched]
            non_zero  = [p for p in probs if p > 0]
            prob_stats = dict(
                min=min(probs) if probs else 0.0,
                max=max(probs) if probs else 0.0,
                non_zero=len(non_zero),
                zero=len(probs) - len(non_zero),
            )

        print(f'\n{sep}')
        print(f'  EVENT: {name:<10}  sheet: {sheet_name:<22}  prob_col: {prob_col}')
        print(sep)
        print(f'  Sheet rows (data) : {total_rows}')
        print(f'  Matched to Core   : {len(matched)}')
        print(f'  Skipped (no Core) : {len(skipped)}')

        if skipped:
            limit = min(len(skipped), 25)
            print(f'  Skipped keys ({limit} of {len(skipped)} shown):')
            for sk in skipped[:limit]:
                print(f'      {sk}')
            if len(skipped) > limit:
                print(f'      ... and {len(skipped) - limit} more')

        if size_dist:
            dist_str = '  '.join(f'{sz}={cnt}' for sz, cnt in sorted(size_dist.items()))
            print(f'  Size distribution : {dist_str}')

        if prob_stats:
            print(f'  Prob col {prob_col} stats  : '
                  f'min={prob_stats["min"]:.4e}  '
                  f'max={prob_stats["max"]:.4e}  '
                  f'non-zero={prob_stats["non_zero"]}  '
                  f'zero={prob_stats["zero"]}')

    print(f'\n{wide}')
    print('SECTION A COMPLETE')
    print(wide + '\n')


# ── Section B ─────────────────────────────────────────────────────────────────

def _compute_impact(ev, sc):
    """Run the impact formula for one scenario. Returns (QY,QX) array or None."""
    formula = ev['formula']
    is_cve  = ev.get('cve', False)

    if is_cve:
        if sc.get('ign_x') is None or sc.get('ign_y') is None:
            return None
        sx, sy = sc['ign_x'], sc['ign_y']
    else:
        sx, sy = sc['sx'], sc['sy']

    d = dist_grid(sx, sy)

    if formula == 'ff':
        return formula_ff(d, sc['lfl_dist'], sc['lflf_dist'])
    if formula == 'thermal':
        return formula_thermal(d, sc['therm_dists'])
    if formula == 'explosion':
        return formula_explosion(d, sc['exp_dists'])
    if formula == 'toxic':
        return formula_toxic(d, sc['tox_dists'], sc['tox_probs'])
    if formula == 'jf_ellipse':
        return formula_jf(sx, sy, sc['dist_vals'], sc['halfW_vals'], sc['center_vals'])
    return None


def run_section_b(wb, core, events):
    wide = '═' * 72
    print(f'\n{wide}')
    print('SECTION B — PER-SCENARIO CONTRIBUTION LOG')
    print(wide)

    header = ['Event', 'Key', 'Size', 'Prob', 'MaxImpact', 'SumImpact', 'NonZeroCells']
    rows   = []

    fmt_h = '{:<10} {:<50} {:<6} {:>12} {:>12} {:>14} {:>13}'
    fmt_r = '{:<10} {:<50} {:<6} {:>12.4e} {:>12.4e} {:>14.4e} {:>13d}'
    print(fmt_h.format(*header))
    print('─' * 120)

    for ev in events:
        name     = ev['name']
        formula  = ev['formula']
        prob_col = ev.get('prob_col')
        reader   = _READER.get(formula)

        if reader is None:
            print(f'[{name}]  formula="{formula}" has no reader — skipping')
            continue

        scenarios = reader(wb, core)

        for sc in scenarios:
            prob   = sc['probs'].get(prob_col, 0.0) if prob_col else 0.0
            impact = _compute_impact(ev, sc)
            if impact is None:
                continue

            max_imp  = float(impact.max())
            sum_imp  = float(impact.sum())
            nz_cells = int(np.count_nonzero(impact))

            row = [name, sc['key'], sc['size'], prob, max_imp, sum_imp, nz_cells]
            rows.append(row)
            print(fmt_r.format(name, sc['key'][:50], sc['size'],
                               prob, max_imp, sum_imp, nz_cells))

    # Save CSV
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    csv_path = os.path.join(OUTPUT_DIR, 'scenario_audit_B.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)

    print(f'\n{wide}')
    print(f'SECTION B COMPLETE — {len(rows)} scenario rows')
    print(f'CSV saved: {csv_path}')
    print(wide + '\n')


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    events = [ev for ev in EVENTS
              if not EVENTS_FILTER or ev['name'] in EVENTS_FILTER]

    print(f'Loading workbook: {EXCEL_PATH}')
    wb   = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    core = read_core(wb)
    print(f'Core: {len(core)} scenarios loaded.\n')

    run_section_a(wb, core, events)

    if RUN_CONTRIBUTION_LOG:
        run_section_b(wb, core, events)
    else:
        print('Section B skipped (RUN_CONTRIBUTION_LOG = False).')
        print('Set RUN_CONTRIBUTION_LOG = True at the top of this script to enable it.\n')


if __name__ == '__main__':
    main()
