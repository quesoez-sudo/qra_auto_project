import openpyxl

# Read with data_only to get cached computed values
wb = openpyxl.load_workbook("KernelV0_v2_copy.xlsx", data_only=True)
ws = wb["ImpactToxMatrix"]

print("=== Key parameter cells (cached values) ===")
for label, cell in [("AA1 (mode)", "AA1"), ("AA2 (risk flag)", "AA2"),
                    ("AA3 (key)", "AA3"), ("AA4 (day/night)", "AA4"),
                    ("AA5 (src_x)", "AA5"), ("AA6 (src_y)", "AA6"),
                    ("AA7 (freq)", "AA7"), ("AA13 (SX)", "AA13"),
                    ("AA14 (SY)", "AA14"), ("AA15 (QX)", "AA15"),
                    ("AA16 (QY)", "AA16"), ("AA17 (limit)", "AA17")]:
    print(f"  {label}: {ws[cell].value}")

# Decode distM for a specific cell to verify units
# Cell at column AB (col index 28 in sheet = grid col 1), row 1 (grid row 1, top-left)
# grid col = COLUMN(n) - col_offset... actually in the ANCHOR formula COLUMN(n)
# starts from 1 within the MAP range so it's 1-based within the area

src_x = ws["AA5"].value
src_y = ws["AA6"].value
sx = ws["AA13"].value
sy = ws["AA14"].value
qx = ws["AA15"].value
qy = ws["AA16"].value

print(f"\n=== Computed distances for corner cells (if src in same units as grid) ===")
print(f"  SX={sx}, SY={sy}, QX={qx}, QY={qy}")
print(f"  src_x={src_x}, src_y={src_y}")

import math
for col_i, row_i in [(1,1),(157,158),(315,317)]:
    x = sx * (col_i - 0.5)
    y = sy * ((qy - row_i) + 0.5)
    dist = math.sqrt((x - src_x)**2 + (y - src_y)**2)
    print(f"  grid col={col_i}, row={row_i} → x={x:.3f}, y={y:.3f}, dist={dist:.4f}")

# Now check J1 and Jmin to understand units
j1 = ws["J1"].value
j_min = None
for row in range(1, 200):
    v = ws.cell(row=row, column=10).value
    if v is None:
        break
    j_min = v
j_last_row = row - 1
print(f"\n=== J column stats ===")
print(f"  J1 (max dist after filter): {j1}")
print(f"  J{j_last_row} (min dist after filter): {j_min}")
print(f"  Conclusion: if src units match J units, distances should be comparable")

# Check a few M values (probability at max and min distance)
m1 = ws["M1"].value
m_last = ws.cell(row=j_last_row, column=13).value
print(f"\n=== M column (prob of fatality) at endpoints ===")
print(f"  M1 (prob at dist={j1}): {m1}")
print(f"  M{j_last_row} (prob at dist={j_min}): {m_last}")

# Check actual non-zero region in AB:MD matrix to confirm
print("\n=== Non-zero cells in matrix (first 5 found) ===")
count = 0
for row in range(1, 318):
    for col in range(28, 343):
        v = ws.cell(row=row, column=col).value
        if v and v > 0:
            # convert to grid coords
            grid_col = col - 27   # 1-based
            grid_row = row
            x = sx * (grid_col - 0.5)
            y = sy * ((qy - grid_row) + 0.5)
            dist_m = math.sqrt((x - src_x)**2 + (y - src_y)**2)
            print(f"  sheet col={col}(grid_col={grid_col}), row={row}(grid_row={grid_row}) → x={x:.3f}, y={y:.3f}, dist={dist_m:.4f}, prob={v:.6f}")
            count += 1
            if count >= 5:
                break
    if count >= 5:
        break

if count == 0:
    print("  No non-zero cells found in AB1:MD317")

wb.close()
